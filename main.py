# -*- coding: utf-8 -*-
from __future__ import annotations

import os, io
from datetime import date, datetime
from typing import Optional, Tuple, List, Dict, Any

from fastapi import FastAPI, HTTPException, Request, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles

from sqlmodel import SQLModel, Field, Session, select, create_engine
from sqlalchemy import func, text
from sqlalchemy.inspection import inspect as sa_inspect

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ===================== DB =====================
def _normalize_database_url(url: str) -> str:
    if url.startswith("postgres://"):
        return "postgresql+psycopg2://" + url[len("postgres://"):]
    if url.startswith("postgresql://") and "+psycopg2" not in url:
        return "postgresql+psycopg2://" + url[len("postgresql://"):]
    return url

DATABASE_URL = (os.getenv("DATABASE_URL") or "").strip()
if DATABASE_URL:
    DATABASE_URL = _normalize_database_url(DATABASE_URL)
    engine = create_engine(DATABASE_URL, pool_pre_ping=True, echo=False)
    DIALECT = "postgres"
else:
    DB_PATH = os.getenv("DB_PATH", "/tmp/maintenance_v3.db")
    os.makedirs(os.path.dirname(DB_PATH) or ".", exist_ok=True)
    engine = create_engine(
        f"sqlite:///{DB_PATH}",
        connect_args={"check_same_thread": False},
        pool_pre_ping=True,
        echo=False,
    )
    DIALECT = "sqlite"

# =================== Models ===================
class Issue(SQLModel, table=True):
    id: Optional[int] = Field(default=None, primary_key=True)
    item_name: str
    model: Optional[str] = None
    serial: Optional[str] = None
    status: Optional[str] = None
    quantity: int = 1
    location: Optional[str] = None
    requester: Optional[str] = None
    issue_date: date
    qualified_by: Optional[str] = None
    receiver: Optional[str] = None

class CabinetRehab(SQLModel, table=True):
    id: Optional[int] = Field(default=None, primary_key=True)
    cabinet_type: str
    code: Optional[str] = Field(default=None, index=True)
    rehab_date: date
    qualified_by: Optional[str] = None
    location: Optional[str] = None
    receiver: Optional[str] = None
    issue_date: Optional[date] = None
    notes: Optional[str] = None

class AssetRehab(SQLModel, table=True):
    id: Optional[int] = Field(default=None, primary_key=True)
    asset_type: str
    model: Optional[str] = None
    serial_or_code: Optional[str] = Field(default=None, index=True)
    quantity: int = 1
    prev_location: Optional[str] = None
    supply_date: date
    qualified_by: Optional[str] = None
    lifted: Optional[bool] = None
    inspector: Optional[str] = None
    tested: Optional[bool] = None
    issue_date: Optional[date] = None
    current_location: Optional[str] = None
    requester: Optional[str] = None
    receiver: Optional[str] = None
    notes: Optional[str] = None
    rehab_date: Optional[date] = Field(default=None, index=True)  # تعتمد عليها الإحصاءات/التقارير

class SparePartRehab(SQLModel, table=True):
    id: Optional[int] = Field(default=None, primary_key=True)
    part_category: str
    part_name: Optional[str] = None
    part_model: Optional[str] = None
    quantity: int = 1
    serial: Optional[str] = Field(default=None, index=True)
    source: Optional[str] = None
    qualified_by: Optional[str] = None
    rehab_date: date
    tested: Optional[bool] = None
    notes: Optional[str] = None

def init_db():
    SQLModel.metadata.create_all(engine)
init_db()

# ---- مهايئ عام: تأكد من وجود العمود rehab_date في assetrehab (SQLite/PG) ----
def ensure_assetrehab_rehab_date():
    insp = sa_inspect(engine)
    try:
        cols = [c["name"] for c in insp.get_columns("assetrehab")]
    except Exception:
        # لو الجدول غير موجود بعد، create_all سيُنشئه؛ لا شيء نفعله هنا.
        return
    if "rehab_date" not in cols:
        ddl = "ALTER TABLE assetrehab ADD COLUMN rehab_date DATE"
        idx = "CREATE INDEX IF NOT EXISTS ix_assetrehab_rehab_date ON assetrehab (rehab_date)"
        with engine.begin() as conn:
            conn.execute(text(ddl))
            conn.execute(text(idx))

# ===================== App ====================
app = FastAPI(title="Maintenance Tracker")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], allow_credentials=True, allow_methods=["*"], allow_headers=["*"]
)
app.mount("/static", StaticFiles(directory="static"), name="static")

@app.on_event("startup")
def _startup_migrate():
    ensure_assetrehab_rehab_date()

@app.get("/")
def root():
    return FileResponse("static/index.html")

@app.get("/healthz")
def healthz():
    return {"ok": True}

# ================ Helpers =====================
def norm(s: Optional[str]) -> Optional[str]:
    if s is None: return None
    t = str(s).strip()
    return t or None

def to_int(x: Optional[str], default=0) -> int:
    try: return int(str(x))
    except: return default

def to_bool(x: Optional[str]) -> Optional[bool]:
    if x is None or x == "": return None
    return str(x).strip().lower() in ("1","true","yes","y","on","نعم")

def to_date(x: Optional[str]) -> Optional[date]:
    if not x: return None
    return datetime.strptime(x, "%Y-%m-%d").date()

def month_bounds(y: int, m: int) -> Tuple[date, date]:
    start = date(y, m, 1)
    end = date(y+1, 1, 1) if m == 12 else date(y, m+1, 1)
    return start, end

def wb_stream(wb: Workbook, filename: str) -> StreamingResponse:
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

def style_header(ws, cols: int, row: int = 1):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="BFE3FF")
        cell.alignment = Alignment(horizontal="center", vertical="center")

def border_all(ws, cols: int):
    thin = Side(style="thin", color="999999")
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=cols):
        for c in r:
            c.border = Border(top=thin, left=thin, right=thin, bottom=thin)

def year_eq(col, y: int):
    return func.strftime("%Y", col) == f"{y:04d}" if DIALECT == "sqlite" else func.extract("year", col) == y

def month_eq(col, m: int):
    return func.strftime("%m", col) == f"{m:02d}" if DIALECT == "sqlite" else func.extract("month", col) == m

# ==================== Issue ===================
@app.post("/api/issue")
async def add_issue(req: Request):
    f = await req.form()
    item = Issue(
        item_name   = f.get("item_name") or "",
        model       = norm(f.get("model")),
        serial      = norm(f.get("serial")),
        status      = norm(f.get("status")),
        quantity    = to_int(f.get("quantity") or "1", 1),
        location    = norm(f.get("location")),
        requester   = norm(f.get("requester")),
        issue_date  = to_date(f.get("issue_date")) or date.today(),
        qualified_by= norm(f.get("qualified_by")),
        receiver    = norm(f.get("receiver")),
    )
    with Session(engine) as s:
        s.add(item); s.commit(); s.refresh(item); return item

@app.get("/api/export/issue/full.xlsx")
def export_issue_full(year: Optional[int] = None, month: Optional[int] = None):
    with Session(engine) as s:
        rows = s.exec(select(Issue).order_by(Issue.issue_date, Issue.id)).all()
    if year and month:
        start, end = month_bounds(year, month)
        rows = [r for r in rows if r.issue_date and start <= r.issue_date < end]
    wb = Workbook(); ws = wb.active; ws.title = "الصرف"; ws.sheet_view.rightToLeft = True
    headers = ["اسم القطعة","المودل","الرقم التسلسلي","الحالة","العدد","الموقع","جهة الطلب","تاريخ الصرف","المؤهل","المستلم"]
    ws.append(headers); style_header(ws, len(headers))
    for r in rows:
        ws.append([r.item_name, r.model, r.serial, r.status, r.quantity, r.location, r.requester, r.issue_date, r.qualified_by, r.receiver])
    border_all(ws, len(headers))
    return wb_stream(wb, f"issue_full{f'_{year}_{month:02d}' if year and month else ''}.xlsx")

@app.get("/api/export/issue/summary.xlsx")
def export_issue_summary(year: Optional[int] = None, month: Optional[int] = None):
    with Session(engine) as s:
        rows = s.exec(select(Issue).order_by(Issue.issue_date, Issue.id)).all()
    if year and month:
        start, end = month_bounds(year, month)
        rows = [r for r in rows if r.issue_date and start <= r.issue_date < end]
    wb = Workbook(); ws = wb.active; ws.title = "ملخص الصرف"; ws.sheet_view.rightToLeft = True
    headers = ["اسم القطعة","العدد","الرقم التسلسلي","الموقع الحالي","المستلم"]
    ws.append(headers); style_header(ws, len(headers))
    for r in rows:
        ws.append([r.item_name, r.quantity, r.serial, r.location, r.receiver])
    border_all(ws, len(headers))
    return wb_stream(wb, f"issue_summary{f'_{year}_{month:02d}' if year and month else ''}.xlsx")

# ================== Cabinets ==================
@app.post("/api/cabinets")
async def add_cabinet(req: Request):
    f = await req.form()
    code = norm(f.get("code"))
    if code:
        with Session(engine) as s:
            dup = s.exec(select(CabinetRehab).where(CabinetRehab.code == code)).first()
            if dup: raise HTTPException(400, "الترميز موجود مسبقًا")
    item = CabinetRehab(
        cabinet_type = f.get("cabinet_type") or "",
        code         = code,
        rehab_date   = to_date(f.get("rehab_date")) or date.today(),
        qualified_by = norm(f.get("qualified_by")),
        location     = norm(f.get("location")),
        receiver     = norm(f.get("receiver")),
        issue_date   = to_date(f.get("issue_date")),
        notes        = norm(f.get("notes")),
    )
    with Session(engine) as s:
        s.add(item); s.commit(); s.refresh(item); return item

@app.get("/api/stats/cabinets")
def stats_cabinets(year: int, month: int):
    cats = ["ATS","AMF","HYBRID","حماية انفرتر","ظفيرة تحكم"]
    res = {k:0 for k in cats}
    with Session(engine) as s:
        q = select(CabinetRehab).where(year_eq(CabinetRehab.rehab_date, year),
                                       month_eq(CabinetRehab.rehab_date, month))
        for r in s.exec(q).all():
            if r.cabinet_type in res: res[r.cabinet_type] += 1
    return res

@app.get("/api/export/cabinets.xlsx")
def export_cabinets(year: int, month: int):
    with Session(engine) as s:
        q = select(CabinetRehab).where(year_eq(CabinetRehab.rehab_date, year),
                                       month_eq(CabinetRehab.rehab_date, month)) \
                                .order_by(CabinetRehab.rehab_date, CabinetRehab.id)
        rows = s.exec(q).all()
    wb = Workbook(); ws = wb.active; ws.title = "الكبائن"; ws.sheet_view.rightToLeft = True
    headers = ["نوع الكبينة","الترميز","تاريخ التأهيل","المؤهل","الموقع","المستلم","تاريخ الصرف","ملاحظات"]
    ws.append(headers); style_header(ws, len(headers))
    for r in rows:
        ws.append([r.cabinet_type, r.code, r.rehab_date, r.qualified_by, r.location, r.receiver, r.issue_date, r.notes])
    border_all(ws, len(headers))
    return wb_stream(wb, f"cabinets_{year}_{month:02d}.xlsx")

# ==================== Assets ==================
def _coerce_asset_payload(d: Dict[str, Any]) -> AssetRehab:
    return AssetRehab(
        asset_type       = d.get("asset_type") or "",
        model            = norm(d.get("model")),
        serial_or_code   = norm(d.get("serial_or_code")),
        quantity         = to_int(d.get("quantity") or "1", 1),
        prev_location    = norm(d.get("prev_location")),
        supply_date      = to_date(d.get("supply_date")) or date.today(),
        qualified_by     = norm(d.get("qualified_by")),
        lifted           = to_bool(d.get("lifted")),
        inspector        = norm(d.get("inspector")),
        tested           = to_bool(d.get("tested")),
        issue_date       = to_date(d.get("issue_date")),
        current_location = norm(d.get("current_location")),
        requester        = norm(d.get("requester")),
        receiver         = norm(d.get("receiver")),
        notes            = norm(d.get("notes")),
        rehab_date       = to_date(d.get("rehab_date")),
    )

def _asset_duplicate_exists(s: Session, serial_or_code: Optional[str], exclude_id: Optional[int] = None) -> bool:
    if not serial_or_code:
        return False
    q = select(AssetRehab.id).where(AssetRehab.serial_or_code == serial_or_code)
    for (rid,) in s.exec(q).all():
        if exclude_id and rid == exclude_id:
            continue
        return True
    return False

@app.post("/api/assets")
async def add_asset(req: Request):
    # تأكد من العمود في أي قاعدة قبل أي عملية
    ensure_assetrehab_rehab_date()

    if "application/json" in (req.headers.get("content-type") or ""):
        f = await req.json()
    else:
        f = dict(await req.form())

    item = _coerce_asset_payload(f)

    with Session(engine) as s:
        if item.serial_or_code and _asset_duplicate_exists(s, item.serial_or_code):
            raise HTTPException(400, "هناك تكرار في الرقم التسلسلي/الترميز")
        s.add(item); s.commit(); s.refresh(item); return item

@app.put("/api/assets/{aid}")
async def update_asset(aid: int, req: Request):
    ensure_assetrehab_rehab_date()

    if (req.headers.get("content-type") or "").lower().startswith("application/json"):
        data = await req.json()
    else:
        form = await req.form()
        data = {k: v for k, v in form.items()}

    with Session(engine) as s:
        obj = s.get(AssetRehab, aid)
        if not obj: raise HTTPException(404, "غير موجود")
        new_serial = norm(data.get("serial_or_code"))
        if new_serial and new_serial != obj.serial_or_code:
            if _asset_duplicate_exists(s, new_serial, exclude_id=aid):
                raise HTTPException(400, "هناك تكرار في الرقم التسلسلي/الترميز")
        patch = _coerce_asset_payload(data).dict()
        patch.pop("id", None)
        for k, v in patch.items():
            setattr(obj, k, v)
        s.add(obj); s.commit(); s.refresh(obj); return obj

@app.get("/api/stats/assets")
def stats_assets(
    year: int = Query(..., description="السنة"),
    month: int = Query(..., ge=1, le=12, description="الشهر 1..12"),
    date_field: str = Query("rehab_date", description="rehab_date أو supply_date")
):
    ensure_assetrehab_rehab_date()

    # افتراضي: rehab_date؛ ويمكن التبديل إلى supply_date
    if date_field == "supply_date":
        col = AssetRehab.supply_date
    else:
        col = func.coalesce(AssetRehab.rehab_date, AssetRehab.supply_date)

    with Session(engine) as s:
        rows = s.exec(
            select(AssetRehab.asset_type, func.coalesce(func.sum(AssetRehab.quantity), 0))
            .where(
                col.is_not(None),
                year_eq(col, year),
                month_eq(col, month),
            )
            .group_by(AssetRehab.asset_type)
        ).all()

    counts = {k: int(v or 0) for (k, v) in rows}
    return {
        "بطاريات":   counts.get("بطاريات", 0),
        "موحدات":    counts.get("موحدات", 0),
        "محركات":    counts.get("محركات", 0),
        "مولدات":    counts.get("مولدات", 0),
        "مكيفات":    counts.get("مكيفات", 0),
        "أصول أخرى": counts.get("أصول أخرى", 0),
    }

def _asset_in_month_with_fallback(r: AssetRehab, start: date, end: date) -> bool:
    d = r.rehab_date or r.supply_date
    return bool(d and start <= d < end)

@app.get("/api/export/assets.xlsx")
def export_assets(year: int, month: int):
    ensure_assetrehab_rehab_date()

    start, end = month_bounds(year, month)
    with Session(engine) as s:
        allrows = s.exec(select(AssetRehab)).all()
        rows = [r for r in allrows if _asset_in_month_with_fallback(r, start, end)]

    wb = Workbook(); ws = wb.active; ws.title = "الأصول"; ws.sheet_view.rightToLeft = True
    headers = ["نوع الأصل","المودل","الرقم التسلسلي/الترميز","العدد","الموقع السابق","تاريخ التوريد",
               "المؤهل","الرفع","الفاحص","الفحص","تاريخ الصرف","الموقع الحالي","جهة الطلب","المستلم","ملاحظات","تاريخ التأهيل"]
    ws.append(headers); style_header(ws, len(headers))
    for r in rows:
        ws.append([
            r.asset_type, r.model, r.serial_or_code, r.quantity, r.prev_location, r.supply_date,
            r.qualified_by, r.lifted, r.inspector, r.tested, r.issue_date, r.current_location,
            r.requester, r.receiver, r.notes, r.rehab_date
        ])
    border_all(ws, len(headers))
    return wb_stream(wb, f"assets_{year}_{month:02d}.xlsx")

# ==================== Spares ==================
@app.post("/api/spares")
async def add_spare(req: Request):
    f = await req.form()
    item = SparePartRehab(
        part_category = f.get("part_category") or "",
        part_name     = norm(f.get("part_name")),
        part_model    = norm(f.get("part_model")),
        quantity      = to_int(f.get("quantity") or "1", 1),
        serial        = norm(f.get("serial")),
        source        = norm(f.get("source")),
        qualified_by  = norm(f.get("qualified_by")),
        rehab_date    = to_date(f.get("rehab_date")) or date.today(),
        tested        = to_bool(f.get("tested")),
        notes         = norm(f.get("notes")),
    )
    with Session(engine) as s:
        s.add(item); s.commit(); s.refresh(item); return item

@app.get("/api/stats/spares")
def stats_spares(year: int, month: int):
    with Session(engine) as s:
        rows = s.exec(
            select(SparePartRehab.part_category, func.coalesce(func.sum(SparePartRehab.quantity), 0))
            .where(year_eq(SparePartRehab.rehab_date, year), month_eq(SparePartRehab.rehab_date, month))
            .group_by(SparePartRehab.part_category)
        ).all()
    cats = ["مضخات الديزل","النوزلات","سلف","دينمو شحن","كروت وشواحن","موديولات","منظمات وانفرترات","تسييخ","أخرى"]
    res = {k:0 for k in cats}
    for k, v in rows:
        if k in res: res[k] = int(v or 0)
    return res

@app.get("/api/export/spares.xlsx")
def export_spares(year: int, month: int):
    with Session(engine) as s:
        q = select(SparePartRehab).where(year_eq(SparePartRehab.rehab_date, year),
                                         month_eq(SparePartRehab.rehab_date, month)) \
                                  .order_by(SparePartRehab.rehab_date, SparePartRehab.id)
        rows = s.exec(q).all()
    wb = Workbook(); ws = wb.active; ws.title = "قطع الغيار"; ws.sheet_view.rightToLeft = True
    headers = ["نوع القطعة","اسم القطعة","موديل القطعة","العدد","الرقم التسلسلي","المصدر","المؤهل","تاريخ التأهيل","الفحص","ملاحظات"]
    ws.append(headers); style_header(ws, len(headers))
    for r in rows:
        ws.append([r.part_category, r.part_name, r.part_model, r.quantity, r.serial, r.source, r.qualified_by, r.rehab_date, r.tested, r.notes])
    border_all(ws, len(headers))
    return wb_stream(wb, f"spares_{year}_{month:02d}.xlsx")

# ============ Duplicates validator ============
@app.get("/api/validate/duplicates")
def validate_duplicates():
    ensure_assetrehab_rehab_date()
    with Session(engine) as s:
        code_counts: Dict[str,int] = {}
        for r in s.exec(select(CabinetRehab)).all():
            if r.code:
                code_counts[r.code] = code_counts.get(r.code, 0) + 1
        cabinets_codes = [k for k,v in code_counts.items() if v > 1]

        ser_counts: Dict[str,int] = {}
        ser_loc_counts: Dict[Tuple[str,str],int] = {}
        for r in s.exec(select(AssetRehab)).all():
            sn = r.serial_or_code or ""
            if sn:
                ser_counts[sn] = ser_counts.get(sn, 0) + 1
                key = (sn, r.current_location or "")
                ser_loc_counts[key] = ser_loc_counts.get(key, 0) + 1
        assets_serials = [k for k,v in ser_counts.items() if v > 1]
        assets_serial_loc_pairs = [f"{a}@{b}" for (a,b),v in ser_loc_counts.items() if v > 1]

        ss: Dict[Tuple[str,str],int] = {}
        for r in s.exec(select(SparePartRehab)).all():
            key = (r.serial or "", r.source or "")
            if key != ("",""):
                ss[key] = ss.get(key, 0) + 1
        spares_serial_src_pairs = [f"{a}@{b}" for (a,b),v in ss.items() if v > 1]

    return {
        "cabinets_codes": cabinets_codes,
        "assets_serials": assets_serials,
        "assets_serial_loc_pairs": assets_serial_loc_pairs,
        "spares_serial_src_pairs": spares_serial_src_pairs
    }

# ======= Monthly & Quarterly summaries ========
AR_MONTHS = ["يناير","فبراير","مارس","أبريل","مايو","يونيو","يوليو","أغسطس","سبتمبر","أكتوبر","نوفمبر","ديسمبر"]

@app.get("/api/export/monthly_summary.xlsx")
def export_monthly_summary(year: int, month: int):
    ensure_assetrehab_rehab_date()

    with Session(engine) as s:
        cab_rows = s.exec(
            select(CabinetRehab.cabinet_type, func.count(CabinetRehab.id))
            .where(year_eq(CabinetRehab.rehab_date, year), month_eq(CabinetRehab.rehab_date, month))
            .group_by(CabinetRehab.cabinet_type)
        ).all()
        cab = {"ATS":0,"AMF":0,"HYBRID":0,"حماية انفرتر":0,"ظفيرة تحكم":0}
        for k, v in cab_rows:
            if k in cab: cab[k] = int(v or 0)

        ast_rows = s.exec(
            select(AssetRehab.asset_type, func.coalesce(func.sum(AssetRehab.quantity), 0))
            .where(
                func.coalesce(AssetRehab.rehab_date, AssetRehab.supply_date).is_not(None),
                year_eq(func.coalesce(AssetRehab.rehab_date, AssetRehab.supply_date), year),
                month_eq(func.coalesce(AssetRehab.rehab_date, AssetRehab.supply_date), month),
            )
            .group_by(AssetRehab.asset_type)
        ).all()
        ast = {"بطاريات":0,"موحدات":0,"محركات":0,"مولدات":0,"مكيفات":0,"أصول أخرى":0}
        for k, v in ast_rows:
            if k in ast: ast[k] = int(v or 0)

        spa_rows = s.exec(
            select(SparePartRehab.part_category, func.coalesce(func.sum(SparePartRehab.quantity), 0))
            .where(year_eq(SparePartRehab.rehab_date, year), month_eq(SparePartRehab.rehab_date, month))
            .group_by(SparePartRehab.part_category)
        ).all()
        spa = {"مضخات الديزل":0,"النوزلات":0,"سلف":0,"دينمو شحن":0,"كروت وشواحن":0,"موديولات":0,"منظمات وانفرترات":0,"تسييخ":0,"أخرى":0}
        for k, v in spa_rows:
            if k in spa: spa[k] = int(v or 0)

    rows_map = [
        ("تجميع كبائن تحكم ATS",         ("cab", "ATS")),
        ("تجميع كبائن تحكم ATS HYBRID",  ("cab", "HYBRID")),
        ("تجميع كبائن تحكم AMF",         ("cab", "AMF")),
        ("تجميع ظفائر مولدات",           ("cab", "ظفيرة تحكم")),
        ("تأهيل موحدات",                  ("ast", "موحدات")),
        ("تأهيل بطاريات",                 ("ast", "بطاريات")),
        ("تأهيل محركات",                  ("ast", "محركات")),
        ("تأهيل مولدات",                  ("ast", "مولدات")),
        ("تأهيل مكيفات",                  ("ast", "مكيفات")),
        ("تأهيل أصول أخرى",               ("ast", "أصول أخرى")),
        ("إصلاح موديولات",                ("spa", "موديولات")),
        ("إصلاح دينمو شحن",               ("spa", "دينمو شحن")),
        ("إصلاح سلف مولد",                ("spa", "سلف")),
        ("إصلاح منظمات شمسية وإنفرترات", ("spa", "منظمات وانفرترات")),
        ("إصلاح كروت وشواحن",            ("spa", "كروت وشواحن")),
        ("إصلاح قطع غيار أخرى",           ("spa", "أخرى")),
    ]
    wb = Workbook(); ws = wb.active; ws.title="ملخص شهري"; ws.sheet_view.rightToLeft=True
    mname = AR_MONTHS[month-1]
    ws.merge_cells("A1:E1")
    ws["A1"].value = f"أهم الإنجازات التي تمت في مركز الإصلاحات الفنية خلال شهر {mname} {year} م:"
    ws["A1"].font = Font(bold=True, size=14, color="003366")
    ws["A1"].alignment = Alignment(horizontal="right")

    headers = ["م","الصنف", mname]
    ws.append(headers); style_header(ws, len(headers), row=3)
    total = 0
    r = 4
    for i,(label,(kind,key)) in enumerate(rows_map, start=1):
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=label)
        v = (cab if kind=="cab" else ast if kind=="ast" else spa).get(key, 0)
        ws.cell(row=r, column=3, value=v)
        total += v; r += 1
    ws.cell(row=r, column=2, value="الإجمالي").font = Font(bold=True)
    ws.cell(row=r, column=3, value=total).font = Font(bold=True)
    border_all(ws, 3)
    return wb_stream(wb, f"monthly_{year}_{month:02d}.xlsx")

@app.get("/api/export/quarterly_summary.xlsx")
def export_quarterly_summary(start_year: int, start_month: int):
    ensure_assetrehab_rehab_date()

    months: List[Tuple[int,int]] = []
    y, m = start_year, start_month
    for _ in range(3):
        months.append((y,m)); m += 1
        if m == 13: m = 1; y += 1

    rows_map = [
        ("تجميع كبائن تحكم ATS",         ("cab", "ATS")),
        ("تجميع كبائن تحكم ATS HYBRID",  ("cab", "HYBRID")),
        ("تجميع كبائن تحكم AMF",         ("cab", "AMF")),
        ("تجميع ظفائر مولدات",           ("cab", "ظفيرة تحكم")),
        ("تأهيل موحدات",                  ("ast", "موحدات")),
        ("تأهيل بطاريات",                 ("ast", "بطاريات")),
        ("تأهيل محركات",                  ("ast", "محركات")),
        ("تأهيل مولدات",                  ("ast", "مولدات")),
        ("تأهيل مكيفات",                  ("ast", "مكيفات")),
        ("تأهيل أصول أخرى",               ("ast", "أصول أخرى")),
        ("إصلاح موديولات",                ("spa", "موديولات")),
        ("إصلاح دينمو شحن",               ("spa", "دينمو شحن")),
        ("إصلاح سلف مولد",                ("spa", "سلف")),
        ("إصلاح منظمات شمسية وإنفرترات", ("spa", "منظمات وانفرترات")),
        ("إصلاح كروت وشواحن",            ("spa", "كروت وشواحن")),
        ("إصلاح قطع غيار أخرى",           ("spa", "أخرى")),
    ]

    monthly_counts: List[Dict[str,int]] = []
    with Session(engine) as s:
        for (yy, mm) in months:
            cab_rows = s.exec(
                select(CabinetRehab.cabinet_type, func.count(CabinetRehab.id))
                .where(year_eq(CabinetRehab.rehab_date, yy), month_eq(CabinetRehab.rehab_date, mm))
                .group_by(CabinetRehab.cabinet_type)
            ).all()
            cab = {"ATS":0,"AMF":0,"HYBRID":0,"حماية انفرتر":0,"ظفيرة تحكم":0}
            for k, v in cab_rows:
                if k in cab: cab[k] = int(v or 0)

            ast_rows = s.exec(
                select(AssetRehab.asset_type, func.coalesce(func.sum(AssetRehab.quantity), 0))
                .where(
                    func.coalesce(AssetRehab.rehab_date, AssetRehab.supply_date).is_not(None),
                    year_eq(func.coalesce(AssetRehab.rehab_date, AssetRehab.supply_date), yy),
                    month_eq(func.coalesce(AssetRehab.rehab_date, AssetRehab.supply_date), mm),
                )
                .group_by(AssetRehab.asset_type)
            ).all()
            ast = {"بطاريات":0,"موحدات":0,"محركات":0,"مولدات":0,"مكيفات":0,"أصول أخرى":0}
            for k, v in ast_rows:
                if k in ast: ast[k] = int(v or 0)

            spa_rows = s.exec(
                select(SparePartRehab.part_category, func.coalesce(func.sum(SparePartRehab.quantity), 0))
                .where(year_eq(SparePartRehab.rehab_date, yy), month_eq(SparePartRehab.rehab_date, mm))
                .group_by(SparePartRehab.part_category)
            ).all()
            spa = {"مضخات الديزل":0,"النوزلات":0,"سلف":0,"دينمو شحن":0,"كروت وشواحن":0,"موديولات":0,"منظمات وانفرترات":0,"تسييخ":0,"أخرى":0}
            for k, v in spa_rows:
                if k in spa: spa[k] = int(v or 0)

            cols: Dict[str,int] = {}
            for label,(kind,key) in rows_map:
                cols[label] = (cab if kind=="cab" else ast if kind=="ast" else spa).get(key, 0)
            monthly_counts.append(cols)

    wb = Workbook(); ws = wb.active; ws.title="ملخص ربع سنوي"; ws.sheet_view.rightToLeft=True
    headers = ["م","الصنف"] + [AR_MONTHS[m-1] for (_,m) in months] + ["الربع"]
    ws.append(headers); style_header(ws, len(headers))
    total_per_month = [0,0,0]; grand_total = 0
    r = 2
    for i,(label,_) in enumerate(rows_map, start=1):
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=label)
        row_sum = 0
        for mi in range(3):
            v = monthly_counts[mi][label]
            ws.cell(row=r, column=3+mi, value=v)
            row_sum += v; total_per_month[mi]+=v
        ws.cell(row=r, column=6, value=row_sum)
        grand_total += row_sum
        r += 1
    ws.cell(row=r, column=2, value="الإجمالي").font = Font(bold=True)
    for mi in range(3): ws.cell(row=r, column=3+mi, value=total_per_month[mi]).font = Font(bold=True)
    ws.cell(row=r, column=6, value=grand_total).font = Font(bold=True)
    border_all(ws, len(headers))
    return wb_stream(wb, f"quarterly_{months[0][0]}_{months[0][1]:02d}.xlsx")
